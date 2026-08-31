"""IFS gold-holdings *changes* importer tests.

These tests are meant to go red if parse/skip/upsert/mapping behavior drifts.
They do not touch the live database or gold_reserves.csv.
"""
import tempfile
import unittest
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from database.models import Base, Country, Metric, TimeSeries, UpdateLog
from pipelines.gold_reserve_changes import (
    CHANGES_METRIC,
    XLSX_PATH,
    import_ifs_changes,
    parse_monthly_sheet,
    resolve_iso,
    run_gold_reserve_changes_fetch,
)


def _tiny_workbook(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly"
    ws["A1"] = "Country Lookup Column"
    ws["B1"] = "Country"
    ws["C1"] = "Comments"
    ws["D1"] = datetime(2020, 1, 1)
    ws["E1"] = datetime(2020, 2, 1)
    rows = [
        ("United States", "United States of America", 1.5, -0.25),
        ("Euro Area (EA)", "Euro Area", 9, 9),
        ("Albania", "Albania", 0, ""),
        ("Zootopia", "Zootopia", 3, 3),
        ("Türkiye, Republic of", "Turkey*", 8, 8),
        ("Turkey", "Turkey", 0.5, -1.0),
        ("State Oil Fund of the Republic of Azerbaijan (SOFAZ)", "State Oil Fund of the Republic of Azerbaijan (SOFAZ)", 2, 2),
        ("China, People's Republic of", "China", 4.0, 5.0),
        ("Taiwan Province of China", "Taiwan, China", 0.1, 0.2),
        ("Macao Special Administrative Region, People's Republic of China", "China, P.R.: Macao", 7, 7),
    ]
    for i, (lookup, short, jan, feb) in enumerate(rows, start=2):
        ws.cell(i, 1, lookup)
        ws.cell(i, 2, short)
        ws.cell(i, 3, None)
        ws.cell(i, 4, jan)
        ws.cell(i, 5, feb if feb != "" else None)

    # Annual must not be imported as the GOLD_RESERVE_CHANGES series.
    annual = wb.create_sheet("Annual")
    annual["A1"] = "Country"
    annual["B1"] = "Comments"
    annual["C1"] = 2026
    annual["A2"] = "United States"
    annual["C2"] = 9999
    annual["A3"] = "China, P.R.: Mainland"
    annual["C3"] = 40.12

    wb.save(path)
    wb.close()


class ResolveIsoTests(unittest.TestCase):
    def test_maps_ifs_aliases(self):
        self.assertEqual(resolve_iso("China, P.R.: Mainland"), "CHN")
        self.assertEqual(resolve_iso("China"), "CHN")
        self.assertEqual(resolve_iso("United States"), "USA")
        self.assertEqual(resolve_iso("United States of America"), "USA")
        self.assertEqual(resolve_iso("Taiwan Province of China"), "TWN")
        self.assertEqual(resolve_iso("Taiwan, China"), "TWN")
        self.assertEqual(resolve_iso("Korea, Rep. of"), "KOR")
        self.assertEqual(resolve_iso("Armenia, Rep. of"), "ARM")
        self.assertEqual(resolve_iso("Czech Rep."), "CZE")
        self.assertEqual(resolve_iso("Turkey"), "TUR")
        self.assertEqual(resolve_iso("Poland"), "POL")
        self.assertEqual(resolve_iso("Egypt"), "EGY")

    def test_skips_non_country_turkey_star_and_macao(self):
        self.assertIsNone(resolve_iso("Euro Area (EA)"))
        self.assertIsNone(resolve_iso("Euro Area"))
        self.assertIsNone(resolve_iso("Turkey*"))
        self.assertIsNone(resolve_iso("Türkiye, Republic of"))
        self.assertIsNone(resolve_iso(
            "State Oil Fund of the Republic of Azerbaijan (SOFAZ)"
        ))
        self.assertIsNone(resolve_iso("China, P.R.: Macao"))
        self.assertIsNone(resolve_iso(
            "Macao Special Administrative Region, People's Republic of China"
        ))
        self.assertIsNone(resolve_iso("Zootopia"))


class ParseMonthlySheetTests(unittest.TestCase):
    def test_real_workbook_date_span(self):
        if not XLSX_PATH.exists():
            self.fail(f"expected staged workbook at {XLSX_PATH}")
        parsed = parse_monthly_sheet(XLSX_PATH)
        self.assertEqual(parsed["date_min"], datetime(2002, 1, 1))
        self.assertEqual(parsed["date_max"], datetime(2026, 6, 1))
        self.assertEqual(parsed["month_count"], 294)
        self.assertGreater(len(parsed["records"]), 1000)
        skipped_names = {s["name"] for s in parsed["skipped"]}
        self.assertTrue(
            any("Euro Area" in n for n in skipped_names),
            skipped_names,
        )
        self.assertTrue(
            any("SOFAZ" in n or "Oil Fund" in n for n in skipped_names),
            skipped_names,
        )
        self.assertTrue(
            any("Turkey*" in n for n in skipped_names),
            skipped_names,
        )
        self.assertTrue(
            any("Macao" in n for n in skipped_names),
            skipped_names,
        )
        isos = {r["iso_code"] for r in parsed["records"]}
        self.assertIn("USA", isos)
        self.assertIn("TUR", isos)
        self.assertIn("CHN", isos)
        self.assertIn("TWN", isos)
        self.assertNotIn("MAC", isos)
        turkey_vals = [
            r for r in parsed["records"]
            if r["iso_code"] == "TUR" and r["date"] == datetime(2020, 1, 1)
        ]
        self.assertEqual(len(turkey_vals), 1)

    def test_monthly_not_annual(self):
        """Series comes from Monthly dates (294 months), not Annual years."""
        parsed = parse_monthly_sheet(XLSX_PATH)
        dates = {r["date"] for r in parsed["records"]}
        self.assertTrue(all(d.day == 1 for d in dates))
        years = {d.year for d in dates}
        self.assertEqual(min(years), 2002)
        self.assertEqual(max(years), 2026)
        # Annual sheet only has one column per year (~25). Monthly has 294.
        self.assertEqual(parsed["month_count"], 294)
        self.assertGreater(len(dates), 200)

    def test_china_2026_ytd_matches_monthly_sum(self):
        parsed = parse_monthly_sheet(XLSX_PATH)
        chn_2026 = [
            r for r in parsed["records"]
            if r["iso_code"] == "CHN" and r["date"].year == 2026
        ]
        self.assertEqual(len(chn_2026), 6)
        months = sorted(r["date"].month for r in chn_2026)
        self.assertEqual(months, [1, 2, 3, 4, 5, 6])
        ytd = sum((r["value"] for r in chn_2026), Decimal("0"))
        self.assertAlmostEqual(float(ytd), 40.12, places=2)
        # Not a single Annual 2026 point.
        self.assertFalse(len(chn_2026) == 1 and chn_2026[0]["date"].month == 1)

    def test_keeps_negatives_and_zeros_skips_empty(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "tiny.xlsx"
            _tiny_workbook(path)
            parsed = parse_monthly_sheet(path)
        by_iso = {}
        for rec in parsed["records"]:
            by_iso.setdefault(rec["iso_code"], {})[rec["date"]] = rec["value"]
        self.assertEqual(by_iso["USA"][datetime(2020, 1, 1)], Decimal("1.50000000"))
        self.assertEqual(by_iso["USA"][datetime(2020, 2, 1)], Decimal("-0.25000000"))
        self.assertEqual(by_iso["ALB"][datetime(2020, 1, 1)], Decimal("0.00000000"))
        self.assertNotIn(datetime(2020, 2, 1), by_iso["ALB"])
        self.assertEqual(by_iso["CHN"][datetime(2020, 1, 1)], Decimal("4.00000000"))
        self.assertEqual(by_iso["TWN"][datetime(2020, 1, 1)], Decimal("0.10000000"))
        self.assertNotIn("MAC", by_iso)
        # Annual 9999 must not appear.
        self.assertNotIn(Decimal("9999.00000000"), {
            rec["value"] for rec in parsed["records"]
        })
        skipped = {(s["name"], s["reason"]) for s in parsed["skipped"]}
        self.assertIn(("Euro Area", "not a country series"), skipped)
        self.assertIn(("Zootopia", "unmapped name"), skipped)
        self.assertIn(("Turkey*", "not a country series"), skipped)
        self.assertTrue(
            any(n.endswith("Macao") or "Macao" in n for n, r in skipped
                if r == "not a country series"),
            skipped,
        )


class ImportSkipUnknownIsoTests(unittest.TestCase):
    def setUp(self):
        engine = create_engine("sqlite:///:memory:")
        Base.metadata.create_all(engine)
        self.Session = sessionmaker(bind=engine)
        self.db = self.Session()
        self.db.add(Country(iso_code="USA", name="United States"))
        self.db.add(Country(iso_code="TUR", name="Turkey"))
        self.db.commit()
        self.tmp = tempfile.TemporaryDirectory()
        self.path = Path(self.tmp.name) / "tiny.xlsx"
        _tiny_workbook(self.path)

    def tearDown(self):
        self.db.close()
        self.tmp.cleanup()

    def test_skips_iso3_not_in_countries_and_does_not_invent(self):
        result = run_gold_reserve_changes_fetch(self.db, xlsx_path=self.path)
        self.assertEqual(result["status"], "success")
        self.assertEqual(result["metric_code"], "GOLD_RESERVE_CHANGES")
        # USA: 2 months, TUR: 2 months. CHN/TWN/ALB mapped but not in countries.
        self.assertEqual(result["inserted"], 4)
        self.assertEqual(result["updated"], 0)
        self.assertEqual(result["countries"], 2)
        reasons = {s["reason"] for s in result["skipped_countries"]}
        self.assertIn("iso3 not in countries", reasons)
        skipped_iso = {
            s.get("iso_code") for s in result["skipped_countries"]
            if s.get("reason") == "iso3 not in countries"
        }
        self.assertIn("CHN", skipped_iso)
        self.assertIn("TWN", skipped_iso)
        self.assertIn("ALB", skipped_iso)
        isos_in_db = {c.iso_code for c in self.db.query(Country).all()}
        self.assertEqual(isos_in_db, {"USA", "TUR"})
        metric = self.db.query(Metric).filter_by(code=CHANGES_METRIC["code"]).first()
        self.assertIsNotNone(metric)
        rows = self.db.query(TimeSeries).filter_by(metric_id=metric.id).all()
        self.assertEqual(len(rows), 4)
        logs = self.db.query(UpdateLog).filter_by(pipeline_name="Gold_Reserve_Changes").all()
        self.assertEqual(len(logs), 1)
        self.assertEqual(logs[0].status, "success")

    def test_second_run_updates_not_duplicates(self):
        run_gold_reserve_changes_fetch(self.db, xlsx_path=self.path)
        again = import_ifs_changes(self.db, xlsx_path=self.path)
        self.assertEqual(again["inserted"], 0)
        self.assertEqual(again["updated"], 4)


if __name__ == "__main__":
    unittest.main()
