"""
IFS Gold Reserve Changes Pipeline
---------------------------------
Imports monthly reported *changes* in central bank gold holdings (tonnes)
from the World Gold Council / IMF IFS workbook.

Source workbook (already staged):
  data/incoming/Changes_latest_as_of_Aug2026_IFS.xlsx
Sheets: Main / Disclaimer / Monthly / Annual
This pipeline reads the Monthly sheet only.

Values are period changes in tonnes (can be negative). Not FRED TRESEG.
Not GOLD_RESERVES levels (see pipelines/gold_reserves.py).

ISO3 codes not present in the countries table are skipped; countries are
never inserted by this pipeline.
"""

import logging
from datetime import datetime, date, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from database.models import Metric, TimeSeries, Country, UpdateLog

logger = logging.getLogger(__name__)

XLSX_PATH = (
    Path(__file__).resolve().parent.parent
    / "data"
    / "incoming"
    / "Changes_latest_as_of_Aug2026_IFS.xlsx"
)
MONTHLY_SHEET = "Monthly"
HEADER_LABEL = "Country Lookup Column"
QUANTIZE = Decimal("0.00000001")

CHANGES_METRIC = {
    "code": "GOLD_RESERVE_CHANGES",
    "name": "Central Bank Gold Reserve Changes",
    "category": "gold",
    "unit": "tonnes",
    "source": "WGC_IFS",
    "description": (
        "Monthly reported changes in central bank gold holdings (tonnes). "
        "World Gold Council / IMF IFS, from January 2002."
    ),
}

# IFS Monthly col A/B labels -> ISO3. Mapping is not an insert.
IFS_COUNTRY_MAP = {
    "Afghanistan, Islamic Republic of": "AFG",
    "Afghanistan, Islamic Rep. of": "AFG",
    "Albania": "ALB",
    "Algeria": "DZA",
    "Argentina": "ARG",
    "Armenia, Republic of": "ARM",
    "Armenia, Rep. of": "ARM",
    "Aruba, Kingdom of the Netherlands": "ABW",
    "Aruba": "ABW",
    "Australia": "AUS",
    "Austria": "AUT",
    "Azerbaijan, Republic of": "AZE",
    "Azerbaijan, Rep. of": "AZE",
    "Bahamas, The": "BHS",
    "Bahrain, Kingdom of": "BHR",
    "Bangladesh": "BGD",
    "Barbados": "BRB",
    "Belarus, Republic of": "BLR",
    "Belarus, Rep. of": "BLR",
    "Belgium": "BEL",
    "Benin": "BEN",
    "Bhutan": "BTN",
    "Bolivia": "BOL",
    "Bosnia and Herzegovina": "BIH",
    "Brazil": "BRA",
    "Brunei Darussalam": "BRN",
    "Bulgaria": "BGR",
    "Burkina Faso": "BFA",
    "Burundi": "BDI",
    "Cabo Verde": "CPV",
    "Cambodia": "KHM",
    "Cameroon": "CMR",
    "Canada": "CAN",
    "Central African Republic": "CAF",
    "Chad": "TCD",
    "Chile": "CHL",
    "Hong Kong Special Administrative Region, People's Republic of China": "HKG",
    "Hong Kong SAR": "HKG",
    "China, People's Republic of": "CHN",
    "China, P.R.: Mainland": "CHN",
    "Colombia": "COL",
    "Comoros, Union of the": "COM",
    "Comoros": "COM",
    "Congo, Democratic Republic of the": "COD",
    "Congo, Democratic Republic of": "COD",
    "Congo, Republic of": "COG",
    "Costa Rica": "CRI",
    "Côte d'Ivoire": "CIV",
    "Cote d'Ivoire": "CIV",
    "Croatia, Republic of": "HRV",
    "Croatia": "HRV",
    "Curaçao and Sint Maarten": "CUW",
    "Curacao & St. Maarten": "CUW",
    "Cyprus": "CYP",
    "Czech Republic": "CZE",
    "Czech Rep.": "CZE",
    "Denmark": "DNK",
    "Dominican Republic": "DOM",
    "Dominican Rep.": "DOM",
    "Ecuador": "ECU",
    "Egypt, Arab Republic of": "EGY",
    "Egypt, Arab Rep. of": "EGY",
    "El Salvador": "SLV",
    "Eritrea, The State of": "ERI",
    "Eritrea": "ERI",
    "Estonia, Republic of": "EST",
    "Estonia, Rep. of": "EST",
    "Ethiopia, The Federal Democratic Republic of": "ETH",
    "Ethiopia, The Federal Dem. Rep. of": "ETH",
    "Fiji, Republic of": "FJI",
    "Fiji": "FJI",
    "Finland": "FIN",
    "France": "FRA",
    "Gabon": "GAB",
    "Georgia": "GEO",
    "Germany": "DEU",
    "Ghana": "GHA",
    "Greece": "GRC",
    "Guatemala": "GTM",
    "Guinea": "GIN",
    "Guinea-Bissau": "GNB",
    "Haiti": "HTI",
    "Honduras": "HND",
    "Hungary": "HUN",
    "Iceland": "ISL",
    "India": "IND",
    "Indonesia": "IDN",
    "Iraq": "IRQ",
    "Ireland": "IRL",
    "Israel": "ISR",
    "Italy": "ITA",
    "Jamaica": "JAM",
    "Japan": "JPN",
    "Jordan": "JOR",
    "Kazakhstan, Republic of": "KAZ",
    "Kazakhstan, Rep. of": "KAZ",
    "Kenya": "KEN",
    "Korea, Republic of": "KOR",
    "Korea, Rep. of": "KOR",
    "Kosovo, Republic of": "XKX",
    "Kuwait": "KWT",
    "Kyrgyz Republic": "KGZ",
    "Kyrgyz Rep.": "KGZ",
    "Lao People's Democratic Republic": "LAO",
    "Latvia, Republic of": "LVA",
    "Latvia": "LVA",
    "Lebanon": "LBN",
    "Libya": "LBY",
    "Lithuania, Republic of": "LTU",
    "Lithuania": "LTU",
    "Luxembourg": "LUX",
    "North Macedonia, Republic of": "MKD",
    "Malawi": "MWI",
    "Malaysia": "MYS",
    "Maldives": "MDV",
    "Mali": "MLI",
    "Malta": "MLT",
    "Mauritania, Islamic Republic of": "MRT",
    "Mauritania": "MRT",
    "Mauritius": "MUS",
    "Mexico": "MEX",
    "Micronesia, Federated States of": "FSM",
    "Moldova, Republic of": "MDA",
    "Moldova": "MDA",
    "Mongolia": "MNG",
    "Montenegro": "MNE",
    "Morocco": "MAR",
    "Mozambique, Republic of": "MOZ",
    "Mozambique": "MOZ",
    "Myanmar": "MMR",
    "Namibia": "NAM",
    "Nepal": "NPL",
    "Netherlands, The": "NLD",
    "Netherlands": "NLD",
    "Netherlands Antilles": "ANT",
    "New Zealand": "NZL",
    "Nicaragua": "NIC",
    "Niger": "NER",
    "Nigeria": "NGA",
    "Norway": "NOR",
    "Oman": "OMN",
    "Pakistan": "PAK",
    "Papua New Guinea": "PNG",
    "Paraguay": "PRY",
    "Peru": "PER",
    "Philippines": "PHL",
    "Poland, Republic of": "POL",
    "Poland, Rep. of": "POL",
    "Portugal": "PRT",
    "Qatar": "QAT",
    "Romania": "ROU",
    "Russian Federation": "RUS",
    "Rwanda": "RWA",
    "San Marino, Republic of": "SMR",
    "San Marino": "SMR",
    "Saudi Arabia": "SAU",
    "Senegal": "SEN",
    "Serbia, Republic of": "SRB",
    "Serbia, Rep. of": "SRB",
    "Singapore": "SGP",
    "Slovak Republic": "SVK",
    "Slovak Rep.": "SVK",
    "Slovenia, Republic of": "SVN",
    "Slovenia, Rep. of": "SVN",
    "Solomon Islands": "SLB",
    "South Africa": "ZAF",
    "South Sudan, Republic of": "SSD",
    "South Sudan": "SSD",
    "Spain": "ESP",
    "Sri Lanka": "LKA",
    "Suriname": "SUR",
    "Sweden": "SWE",
    "Switzerland": "CHE",
    "Syrian Arab Republic": "SYR",
    "Taiwan Province of China": "TWN",
    "Tajikistan, Republic of": "TJK",
    "Tajikistan, Rep. of": "TJK",
    "Thailand": "THA",
    "Timor-Leste, Democratic Republic of": "TLS",
    "Timor-Leste, Dem. Rep. of": "TLS",
    "Togo": "TGO",
    "Trinidad and Tobago": "TTO",
    "Tunisia": "TUN",
    "Turkey": "TUR",
    "Ukraine": "UKR",
    "United Arab Emirates": "ARE",
    "United Kingdom": "GBR",
    "United States": "USA",
    "Uruguay": "URY",
    "Uzbekistan, Republic of": "UZB",
    "Uzbekistan, Rep. of": "UZB",
    "Venezuela, República Bolivariana de": "VEN",
    "Venezuela, Republica Bolivariana de": "VEN",
    "Yemen, Republic of": "YEM",
    "Zambia": "ZMB",
    "Zimbabwe": "ZWE",
}

# Display-name aliases (WGC / colloquial) that differ from IFS Monthly labels.
# Mapper resolves these to ISO3; countries are still never inserted here.
IFS_NAME_ALIASES = {
    "China": "CHN",
    "China, Mainland": "CHN",
    "United States of America": "USA",
    "Taiwan, China": "TWN",
    "Taiwan": "TWN",
    "Egypt": "EGY",
    "Poland": "POL",
    "South Korea": "KOR",
    "Czechia": "CZE",
    "Russia": "RUS",
    "Hong Kong": "HKG",
    "Venezuela": "VEN",
    "Syria": "SYR",
    "Laos": "LAO",
    "Slovakia": "SVK",
    "Kyrgyzstan": "KGZ",
    "North Macedonia": "MKD",
    "Armenia": "ARM",
    "Azerbaijan": "AZE",
    "Belarus": "BLR",
    "Kazakhstan": "KAZ",
    "Uzbekistan": "UZB",
    "Serbia": "SRB",
    "Estonia": "EST",
    "Slovenia": "SVN",
    "Tajikistan": "TJK",
    "Afghanistan": "AFG",
    "Bahrain": "BHR",
    "Ethiopia": "ETH",
    "Kosovo": "XKX",
}

# Euro Area / SOFAZ / Macao have no ISO in WGC_COUNTRY_MAP — do not invent.
# Turkey* is the gross duplicate; official Turkey -> TUR is kept.
SKIP_NAMES = {
    "euro area (ea)",
    "euro area",
    "state oil fund of the republic of azerbaijan (sofaz)",
    "türkiye, republic of",
    "turkiye, republic of",
    "turkey*",
    "macao special administrative region, people's republic of china",
    "china, p.r.: macao",
    "macao",
    "macau",
    "china, p.r.: macau",
}


def _norm_name(name) -> str:
    if name is None:
        return ""
    text = str(name).replace("\xa0", " ").strip()
    return " ".join(text.split())


def resolve_iso(name) -> str | None:
    """Map an IFS country label to ISO3, or None to skip.

    Exact match on Monthly col A/B (IFS_COUNTRY_MAP) or display aliases
    (IFS_NAME_ALIASES). Does not insert countries.
    """
    raw = _norm_name(name)
    if not raw:
        return None
    if raw.casefold() in SKIP_NAMES:
        return None
    if raw in IFS_COUNTRY_MAP:
        return IFS_COUNTRY_MAP[raw]
    if raw in IFS_NAME_ALIASES:
        return IFS_NAME_ALIASES[raw]
    return None


def _to_month_start(value) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return datetime(value.year, value.month, 1)
    if isinstance(value, date):
        return datetime(value.year, value.month, 1)
    try:
        serial = float(value)
    except (TypeError, ValueError):
        return None
    parsed = datetime(1899, 12, 30) + timedelta(days=serial)
    return datetime(parsed.year, parsed.month, 1)


def _to_tonnes(value) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text or text.upper() in {"N/A", "#N/A", "AWAITED", "-", "NA"}:
            return None
        try:
            tonnes = Decimal(text)
        except Exception:
            return None
    else:
        try:
            tonnes = Decimal(str(float(value)))
        except (TypeError, ValueError):
            return None
    return tonnes.quantize(QUANTIZE, rounding=ROUND_HALF_UP)


def ensure_changes_metric(db: Session) -> Metric:
    metric = db.query(Metric).filter_by(code=CHANGES_METRIC["code"]).first()
    if not metric:
        metric = Metric(**CHANGES_METRIC)
        db.add(metric)
        db.commit()
        logger.info("Created metric: GOLD_RESERVE_CHANGES")
    return metric


def parse_monthly_sheet(xlsx_path: Path) -> dict:
    """Parse Monthly sheet into records plus skip notes. No DB writes."""
    if not xlsx_path.exists():
        raise FileNotFoundError(
            f"IFS gold-holdings changes workbook not found at {xlsx_path}"
        )

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if MONTHLY_SHEET not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet {MONTHLY_SHEET!r} not found in {xlsx_path.name}")
    ws = wb[MONTHLY_SHEET]

    header_seen = False
    dates = []
    records = []
    skipped = []
    seen_skip = set()

    def note_skip(name: str, reason: str):
        key = (name, reason)
        if key in seen_skip:
            return
        seen_skip.add(key)
        skipped.append({"name": name, "reason": reason})

    for row in ws.iter_rows(values_only=True):
        cells = list(row)
        if not cells:
            continue
        first = _norm_name(cells[0] if cells else None)
        if not header_seen:
            if first == HEADER_LABEL:
                header_seen = True
                dates = [_to_month_start(v) for v in cells[3:]]
            continue

        lookup = _norm_name(cells[0] if len(cells) > 0 else None)
        short = _norm_name(cells[1] if len(cells) > 1 else None)
        label = short or lookup
        if not label:
            continue

        iso = resolve_iso(short) or resolve_iso(lookup)
        if iso is None:
            folded = {lookup.casefold(), short.casefold(), label.casefold()}
            reason = (
                "not a country series"
                if folded & SKIP_NAMES
                else "unmapped name"
            )
            note_skip(label, reason)
            continue

        for i, raw in enumerate(cells[3:]):
            if i >= len(dates) or dates[i] is None:
                continue
            tonnes = _to_tonnes(raw)
            if tonnes is None:
                continue
            records.append({
                "iso_code": iso,
                "country_name": label,
                "date": dates[i],
                "value": tonnes,
            })

    wb.close()
    valid_dates = [d for d in dates if d is not None]
    return {
        "records": records,
        "skipped": skipped,
        "date_min": min(valid_dates) if valid_dates else None,
        "date_max": max(valid_dates) if valid_dates else None,
        "month_count": len(valid_dates),
    }


def import_ifs_changes(db: Session, xlsx_path: Path = XLSX_PATH) -> dict:
    """Parse the workbook and upsert TimeSeries for GOLD_RESERVE_CHANGES."""
    parsed = parse_monthly_sheet(xlsx_path)
    metric = ensure_changes_metric(db)
    inserted = updated = 0
    skipped_not_in_db = []
    countries_found = set()
    iso_missing = set()
    country_cache = {}

    def get_country(iso_code: str):
        if iso_code in country_cache:
            return country_cache[iso_code]
        country = db.query(Country).filter_by(iso_code=iso_code).first()
        country_cache[iso_code] = country
        return country

    for rec in parsed["records"]:
        iso = rec["iso_code"]
        country = get_country(iso)
        if not country:
            if iso not in iso_missing:
                iso_missing.add(iso)
                skipped_not_in_db.append({
                    "name": rec["country_name"],
                    "iso_code": iso,
                    "reason": "iso3 not in countries",
                })
            continue

        countries_found.add(iso)
        existing = db.query(TimeSeries).filter(
            TimeSeries.metric_id == metric.id,
            TimeSeries.country_id == country.id,
            TimeSeries.date == rec["date"],
        ).first()

        if existing:
            existing.value = rec["value"]
            existing.updated_at = datetime.utcnow()
            updated += 1
        else:
            db.add(TimeSeries(
                metric_id=metric.id,
                country_id=country.id,
                date=rec["date"],
                value=rec["value"],
            ))
            inserted += 1

    db.commit()
    skipped = parsed["skipped"] + skipped_not_in_db
    logger.info(
        "IFS gold changes import: %s inserted, %s updated, "
        "%s countries skipped, %s countries",
        inserted, updated, len(skipped), len(countries_found),
    )
    return {
        "status": "success",
        "metric_code": CHANGES_METRIC["code"],
        "inserted": inserted,
        "updated": updated,
        "skipped": len(skipped),
        "skipped_countries": skipped,
        "countries": len(countries_found),
        "date_min": parsed["date_min"].strftime("%Y-%m") if parsed["date_min"] else None,
        "date_max": parsed["date_max"].strftime("%Y-%m") if parsed["date_max"] else None,
        "month_count": parsed["month_count"],
    }


def run_gold_reserve_changes_fetch(db: Session, xlsx_path: Path = XLSX_PATH) -> dict:
    """Entry point compatible with scheduler / manual trigger / local run."""
    start_time = datetime.utcnow()
    try:
        result = import_ifs_changes(db, xlsx_path=xlsx_path)
        db.add(UpdateLog(
            pipeline_name="Gold_Reserve_Changes",
            status="success",
            records_inserted=result["inserted"],
            records_updated=result["updated"],
            error_message=None,
            started_at=start_time,
            completed_at=datetime.utcnow(),
        ))
        db.commit()
        return result
    except Exception as e:
        logger.error("IFS gold reserve changes import failed: %s", e)
        db.add(UpdateLog(
            pipeline_name="Gold_Reserve_Changes",
            status="failed",
            records_inserted=0,
            records_updated=0,
            error_message=str(e)[:500],
            started_at=start_time,
            completed_at=datetime.utcnow(),
        ))
        db.commit()
        return {
            "status": "failed",
            "metric_code": CHANGES_METRIC["code"],
            "inserted": 0,
            "updated": 0,
            "errors": [str(e)],
        }
